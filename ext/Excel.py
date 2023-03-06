import openpyxl

class Excel(object):
    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.load_workbook(self.path)
        self.ws = self.wb.active

    def get_row_count(self):
        return self.ws.max_row

    def get_column_count(self):
        return self.ws.max_column

    def read_data(self, row, column):
        return self.ws.cell(row=row, column=column).value

    def write_data(self, row, column, value):
        self.ws.cell(row=row, column=column).value = value
        self.wb.save(self.path)